from django.core.management.base import BaseCommand
from django.db import transaction
from item_oil_gas.models import Item
import openpyxl

class Command(BaseCommand):
    help = "Importa items desde un archivo Excel (.xlsx) a Item Oil & Gas"

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Ruta del archivo .xlsx")

    def handle(self, *args, **options):
        path = options["path"]

        wb = openpyxl.load_workbook(path)
        ws = wb[wb.sheetnames[0]]

        raw_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [(h or "").strip() for h in raw_headers]
        col_idx = {headers[i]: i + 1 for i in range(len(headers)) if headers[i]}

        required = ["Código", "Descripción", "Unidad Medida", "Clasificacion", "Grupo Inventario"]
        missing = [h for h in required if h not in col_idx]
        if missing:
            self.stderr.write(self.style.ERROR(f"Faltan columnas: {', '.join(missing)}"))
            return

        creados = 0
        actualizados = 0
        vacios = 0

        with transaction.atomic():
            for r in range(2, ws.max_row + 1):
                codigo = (ws.cell(row=r, column=col_idx["Código"]).value or "").strip()
                if not codigo:
                    vacios += 1
                    continue

                defaults = {
                    "descripcion": (ws.cell(row=r, column=col_idx["Descripción"]).value or "").strip(),
                    "unidad_medida": (ws.cell(row=r, column=col_idx["Unidad Medida"]).value or "").strip(),
                    "clasificacion": (ws.cell(row=r, column=col_idx["Clasificacion"]).value or "").strip(),
                    "grupo_inventario": (ws.cell(row=r, column=col_idx["Grupo Inventario"]).value or "").strip(),
                }

                _, created = Item.objects.update_or_create(codigo=codigo, defaults=defaults)
                if created:
                    creados += 1
                else:
                    actualizados += 1

        self.stdout.write(self.style.SUCCESS(
            f"OK. Creados: {creados} | Actualizados: {actualizados} | Filas sin código: {vacios}"
        ))