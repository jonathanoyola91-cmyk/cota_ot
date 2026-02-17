from django.contrib import messages
from django.db import transaction
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView

from .models import Item
from .forms import ItemForm, ItemImportForm

import openpyxl


class ItemListView(ListView):
    model = Item
    template_name = "item_oil_gas/item_list.html"
    context_object_name = "items"
    paginate_by = 50

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(codigo__icontains=q) | qs.filter(descripcion__icontains=q)
        return qs


class ItemCreateView(CreateView):
    model = Item
    form_class = ItemForm
    template_name = "item_oil_gas/item_form.html"
    success_url = reverse_lazy("item_oil_gas:item_list")


class ItemUpdateView(UpdateView):
    model = Item
    form_class = ItemForm
    template_name = "item_oil_gas/item_form.html"
    success_url = reverse_lazy("item_oil_gas:item_list")


class ItemDetailView(DetailView):
    model = Item
    template_name = "item_oil_gas/item_detail.html"
    context_object_name = "item"


class ItemDeleteView(DeleteView):
    model = Item
    template_name = "item_oil_gas/item_confirm_delete.html"
    success_url = reverse_lazy("item_oil_gas:item_list")


def import_items(request):
    """
    Importa Excel (.xlsx) y hace UPSERT por 'codigo':
    - si existe: actualiza
    - si no existe: crea
    """
    if request.method == "POST":
        form = ItemImportForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[wb.sheetnames[0]]

                # headers en tu archivo vienen con espacios: "Código  ", etc.
                raw_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                headers = [(h or "").strip() for h in raw_headers]

                # Mapa esperado (por tu excel)
                expected = {
                    "Código": "codigo",
                    "Descripción": "descripcion",
                    "Unidad Medida": "unidad_medida",
                    "Clasificacion": "clasificacion",
                    "Grupo Inventario": "grupo_inventario",
                }

                # Crear índice de columnas por header
                col_idx = {headers[i]: i + 1 for i in range(len(headers)) if headers[i]}

                missing = [h for h in expected.keys() if h not in col_idx]
                if missing:
                    messages.error(request, f"Faltan columnas en el Excel: {', '.join(missing)}")
                    return redirect("item_oil_gas:item_import")

                creados = 0
                actualizados = 0
                vacios = 0

                with transaction.atomic():
                    for r in range(2, ws.max_row + 1):
                        codigo = (ws.cell(row=r, column=col_idx["Código"]).value or "").strip()
                        if not codigo:
                            vacios += 1
                            continue

                        data = {
                            "descripcion": (ws.cell(row=r, column=col_idx["Descripción"]).value or "").strip(),
                            "unidad_medida": (ws.cell(row=r, column=col_idx["Unidad Medida"]).value or "").strip(),
                            "clasificacion": (ws.cell(row=r, column=col_idx["Clasificacion"]).value or "").strip(),
                            "grupo_inventario": (ws.cell(row=r, column=col_idx["Grupo Inventario"]).value or "").strip(),
                        }

                        obj, created = Item.objects.update_or_create(
                            codigo=codigo,
                            defaults=data,
                        )
                        if created:
                            creados += 1
                        else:
                            actualizados += 1

                messages.success(
                    request,
                    f"Importación lista. Creados: {creados} | Actualizados: {actualizados} | Filas sin código: {vacios}"
                )
                return redirect("item_oil_gas:item_list")

            except Exception as e:
                messages.error(request, f"Error importando Excel: {e}")
                return redirect("item_oil_gas:item_import")
    else:
        form = ItemImportForm()

    return render(request, "item_oil_gas/item_import.html", {"form": form})

from django.http import HttpResponse
import openpyxl


def download_template(request):
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    # Encabezados EXACTOS (como los espera el importador)
    headers = ["Código", "Descripción", "Unidad Medida", "Clasificacion", "Grupo Inventario"]
    ws.append(headers)

    # Fila ejemplo (opcional)
    ws.append(["EJ-0001", "ITEM DE EJEMPLO", "UND", "REPUESTOS", "GRUPO 1"])

    # Preparar respuesta HTTP como archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_items_oil_gas.xlsx"'

    wb.save(response)
    return response