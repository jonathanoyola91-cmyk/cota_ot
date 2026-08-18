from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.roles import tiene_rol
from .forms import SupplierForm
from .models import PurchaseRequest, PurchaseLine, Supplier


def _get_paw_from_compra(compra):
    """Obtiene el PAW asociado a la solicitud de compra sin romper si falta alguna relación."""
    try:
        return compra.bom.workorder.paw
    except AttributeError:
        return None


def _estado_es_aprobado(obj):
    """Valida estados aprobados soportando texto y constantes de modelos."""
    return str(getattr(obj, "estado", "")).upper() in ["APROBADO", "APROBADA"]


def _tiene_lineas_contado(compra):
    return compra.lineas.filter(cantidad_requerida__gt=0, tipo_pago="CONTADO").exists()


def _tiene_lineas_credito(compra):
    return compra.lineas.filter(cantidad_requerida__gt=0, tipo_pago="CREDITO").exists()


def _finanzas_aprobado(compra):
    try:
        from finanzas.models import FinanceApproval
    except Exception:
        return False

    approvals = FinanceApproval.objects.filter(purchase_request=compra)
    return any(_estado_es_aprobado(aprob) for aprob in approvals)


def _aprobacion_aprobada(compra):
    try:
        from aprobacion.models import PurchaseApproval
    except Exception:
        return False

    approvals = PurchaseApproval.objects.filter(purchase_request=compra)
    return any(_estado_es_aprobado(aprob) for aprob in approvals)




def _estado_flujo_linea(linea):
    """Devuelve el estado operativo de aprobación de una línea de compra.

    CONTADO  -> Finanzas (FinanceApprovalLine)
    CREDITO  -> Gerencia (PurchaseApprovalLine)
    """
    if Decimal(linea.cantidad_a_comprar or 0) <= 0:
        return {"estado": "NO_APLICA", "texto": "No requiere compra", "aprobado": True, "enviado": True}

    if not linea.proveedor_id or linea.precio_unitario is None:
        return {"estado": "INCOMPLETO", "texto": "Completar proveedor y precio", "aprobado": False, "enviado": False}

    if linea.tipo_pago == "CONTADO":
        try:
            fl = linea.finance_line
        except Exception:
            return {"estado": "PENDIENTE_ENVIO", "texto": "Pendiente enviar a Finanzas", "aprobado": False, "enviado": False}

        decision = str(fl.decision or "PENDIENTE").upper()
        if fl.pagado:
            return {"estado": "APROBADO", "texto": "Pagado por Finanzas", "aprobado": True, "enviado": True}
        if decision == "APROBADO":
            return {"estado": "APROBADO", "texto": "Aprobado por Finanzas", "aprobado": True, "enviado": True}
        if decision == "PROGRAMADO":
            return {"estado": "APROBADO", "texto": "Programado por Finanzas", "aprobado": True, "enviado": True}
        if decision == "RECHAZADO":
            return {"estado": "RECHAZADO", "texto": "Rechazado por Finanzas", "aprobado": False, "enviado": True}
        if decision == "EN_ESPERA":
            return {"estado": "EN_ESPERA", "texto": "En espera en Finanzas", "aprobado": False, "enviado": True}
        return {"estado": "PENDIENTE", "texto": "Pendiente en Finanzas", "aprobado": False, "enviado": True}

    if linea.tipo_pago == "CREDITO":
        try:
            al = linea.purchase_approval_line
        except Exception:
            return {"estado": "PENDIENTE_ENVIO", "texto": "Pendiente enviar a Gerencia", "aprobado": False, "enviado": False}

        estado = str(al.estado_aprobacion or "PENDIENTE").upper()
        if estado == "APROBADO":
            return {"estado": "APROBADO", "texto": "Aprobado por Gerencia", "aprobado": True, "enviado": True}
        if estado == "RECHAZADO":
            return {"estado": "RECHAZADO", "texto": "Rechazado por Gerencia", "aprobado": False, "enviado": True}
        return {"estado": "PENDIENTE", "texto": "Pendiente en Gerencia", "aprobado": False, "enviado": True}

    return {"estado": "INCOMPLETO", "texto": "Definir tipo de pago", "aprobado": False, "enviado": False}


def _resumen_aprobaciones(compra):
    lineas = list(
        compra.lineas
        .filter(cantidad_requerida__gt=0, cantidad_a_comprar__gt=0)
        .select_related("proveedor")
    )

    total = len(lineas)
    aprobadas = 0
    contado_total = contado_aprobado = 0
    credito_total = credito_aprobado = 0

    for linea in lineas:
        estado = _estado_flujo_linea(linea)
        if estado["aprobado"]:
            aprobadas += 1
        if linea.tipo_pago == "CONTADO":
            contado_total += 1
            if estado["aprobado"]:
                contado_aprobado += 1
        elif linea.tipo_pago == "CREDITO":
            credito_total += 1
            if estado["aprobado"]:
                credito_aprobado += 1

    return {
        "total": total,
        "aprobadas": aprobadas,
        "contado_total": contado_total,
        "contado_aprobado": contado_aprobado,
        "credito_total": credito_total,
        "credito_aprobado": credito_aprobado,
        "todas_aprobadas": total > 0 and aprobadas == total,
    }


def _recepcion_creada(compra):
    try:
        from inventario.models import InventoryReception
    except Exception:
        return False

    return InventoryReception.objects.filter(purchase_request=compra).exists()


def _linea_en_inventario(linea):
    """Indica si la línea de compra ya fue enviada a Inventario."""
    try:
        from inventario.models import InventoryReceptionLine
    except Exception:
        return False

    return InventoryReceptionLine.objects.filter(purchase_line=linea).exists()


def _recepcion_completa(compra):
    """
    La recepción solo se considera completa cuando existe recepción
    y todas sus líneas reales tienen cantidad recibida completa.
    Esto evita entregar a taller o cerrar compra solo por haber creado
    el registro de recepción.
    """
    try:
        from inventario.models import InventoryReception
    except Exception:
        return False

    recepcion = (
        InventoryReception.objects
        .filter(purchase_request=compra)
        .prefetch_related("lineas")
        .first()
    )

    if not recepcion:
        return False

    lineas = list(recepcion.lineas.all())

    if not lineas:
        return False

    for linea in lineas:
        esperada = Decimal(linea.cantidad_esperada or 0)
        recibida = Decimal(linea.cantidad_recibida or 0)

        if esperada > 0 and recibida < esperada:
            return False

        if str(getattr(linea, "estado", "")).upper() == "PENDIENTE":
            return False

    return True


def _get_entrega(compra):
    try:
        from inventario.models import WorkshopDelivery
    except Exception:
        return None
    return (
        WorkshopDelivery.objects
        .filter(purchase_request=compra)
        .prefetch_related("lineas")
        .first()
    )


def _entrega_completa(compra):
    entrega = _get_entrega(compra)
    if not entrega:
        return False

    lineas = list(entrega.lineas.all())
    if not lineas:
        return False

    for linea in lineas:
        requerida = Decimal(linea.cantidad_requerida or 0)
        entregada = Decimal(linea.cantidad_entregada or 0)
        if requerida > 0 and entregada < requerida:
            return False
    return True

@require_POST
@login_required
def aprobar_gerencia_compra(request, pk):
    if not tiene_rol(request.user, ["GERENTE", "ADMIN"]):
        messages.error(request, "Solo gerencia puede aprobar esta compra.")
        return redirect("/")

    from aprobacion.models import PurchaseApproval

    compra = get_object_or_404(PurchaseRequest, pk=pk)

    aprobacion, created = PurchaseApproval.objects.get_or_create(
        purchase_request=compra,
        defaults={
            "estado": PurchaseApproval.Estado.PENDIENTE,
            "enviado_por": request.user,
        },
    )

    aprobacion.estado = PurchaseApproval.Estado.APROBADO
    aprobacion.aprobado_por = request.user
    aprobacion.save()

    paw = _get_paw_from_compra(compra)
    if paw:
        paw.estado_operativo = "APROBADO_GERENCIA"
        paw.save(update_fields=["estado_operativo"])

    messages.success(request, "Compra aprobada por gerencia. Ya puede enviarse a Inventario.")
    return redirect("compras_oil:paw_detail", pk=compra.pk)

@login_required
def dashboard(request):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes acceso a Compras.")
        return redirect("/")

    compras_all = PurchaseRequest.objects.all().order_by("-actualizado_en")

    compras = (
        PurchaseRequest.objects
        .exclude(estado="CERRADA")
        .select_related("bom", "bom__workorder", "creado_por")
        .annotate(
            # Solo se contabilizan líneas reales: cantidad requerida mayor a cero.
            total_lineas=Count(
                "lineas",
                filter=Q(lineas__cantidad_requerida__gt=0),
                distinct=True,
            ),
            lineas_diligenciadas=Count(
                "lineas",
                filter=(
                    Q(lineas__cantidad_requerida__gt=0) &
                    Q(lineas__proveedor__isnull=False) &
                    Q(lineas__precio_unitario__isnull=False)
                ),
                distinct=True,
            ),
            lineas_pagadas=Count(
                "lineas__finance_line",
                filter=(
                    Q(lineas__cantidad_requerida__gt=0) &
                    Q(lineas__finance_line__pagado=True)
                ),
                distinct=True,
            )
        )
        .order_by("-actualizado_en")
    )

    for compra in compras:
        compra.porcentaje_avance = int(
            (compra.lineas_diligenciadas / compra.total_lineas) * 100
        ) if compra.total_lineas > 0 else 0

    context = {
        "compras": compras,
        "total_solicitudes": compras_all.count(),
        "total_borrador": compras_all.filter(estado="BORRADOR").count(),
        "total_revision": compras_all.filter(estado="EN_REVISION").count(),
        "total_cerrada": compras_all.filter(estado="CERRADA").count(),
        "total_activas": compras.count(),
    }

    return render(request, "compras_oil/dashboard.html", context)


@login_required
def compras_dashboard(request):
    return dashboard(request)


@login_required
def historial_compras(request):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes acceso al historial de Compras.")
        return redirect("/")

    q = (request.GET.get("q") or "").strip()

    compras = (
        PurchaseRequest.objects
        .filter(estado="CERRADA")
        .select_related("bom", "bom__workorder", "creado_por")
        .annotate(
            total_lineas=Count(
                "lineas",
                filter=Q(lineas__cantidad_requerida__gt=0),
                distinct=True,
            ),
            total_proveedores=Count(
                "lineas__proveedor",
                filter=Q(lineas__cantidad_requerida__gt=0),
                distinct=True,
            ),
        )
        .order_by("-actualizado_en")
    )

    if q:
        filtro = (
            Q(paw_numero__icontains=q) |
            Q(paw_nombre__icontains=q) |
            Q(lineas__codigo__icontains=q) |
            Q(lineas__descripcion__icontains=q) |
            Q(lineas__proveedor__nombre__icontains=q)
        )
        if q.isdigit():
            filtro |= Q(bom__workorder__numero=int(q))
        compras = compras.filter(filtro).distinct()

    return render(request, "compras_oil/historial.html", {
        "compras": compras,
        "q": q,
        "total_cerradas": PurchaseRequest.objects.filter(estado="CERRADA").count(),
    })


@require_POST
@login_required
def cerrar_solicitud(request, pk):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para cerrar solicitudes.")
        return redirect("/")

    compra = get_object_or_404(PurchaseRequest, pk=pk)
    paw = _get_paw_from_compra(compra)

    if compra.estado == "CERRADA":
        messages.info(request, "Esta compra ya se encuentra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if not _recepcion_completa(compra):
        messages.error(request, "No puedes cerrar la compra. Primero debes completar la recepción del material.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    entrega = _get_entrega(compra)
    if not entrega:
        messages.error(request, "No puedes cerrar la compra. Primero debes definir y registrar la entrega del material.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if not _entrega_completa(compra):
        messages.error(request, "No puedes cerrar la compra. La entrega todavía tiene cantidades pendientes.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    compra.estado = "CERRADA"
    compra.save(update_fields=["estado", "actualizado_en"])

    # Cerrar la COMPRA no equivale a cerrar todo el PAW.
    # Solo avanzamos el PAW según el destino y su alcance real.
    if paw:
        destino = getattr(entrega, "destino", "TALLER")
        if destino == "TALLER" and getattr(paw, "aplica_taller", True):
            paw.estado_operativo = "ENTREGADO_TALLER"
            paw.save(update_fields=["estado_operativo"])
        elif destino == "INVENTARIO":
            if not getattr(paw, "aplica_taller", False) and not getattr(paw, "aplica_campo", False):
                paw.estado_operativo = "PRODUCTO_OK"
                paw.save(update_fields=["estado_operativo"])
        # CAMPO conserva MATERIAL_RECIBIDO: el módulo Campo continúa el flujo.

    messages.success(request, f"Compra PAW {compra.paw_numero} cerrada correctamente.")
    return redirect("compras_oil:dashboard")

@login_required
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)

    return render(request, "compras_oil/supplier_detail.html", {
        "supplier": supplier,
    })


@login_required
def supplier_create(request):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para crear proveedores.")
        return redirect("/")

    next_url = request.GET.get("next")

    if request.method == "POST":
        form = SupplierForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor creado correctamente.")

            if next_url:
                return redirect(next_url)

            return redirect("compras_oil:dashboard")
    else:
        form = SupplierForm()

    return render(request, "compras_oil/supplier_form.html", {
        "form": form,
        "next_url": next_url,
    })

@login_required
def supplier_list(request):
    if not tiene_rol(request.user, ["COMPRAS", "FINANZAS", "GERENTE", "ADMIN"]):
        messages.error(request, "No tienes acceso a proveedores.")
        return redirect("/")

    suppliers = Supplier.objects.all().order_by("nombre")

    return render(request, "compras_oil/supplier_list.html", {
        "suppliers": suppliers
    })

@login_required
def purchase_request_pdf(request, pk: int):
    pr = get_object_or_404(
        PurchaseRequest.objects
        .select_related("bom", "bom__workorder")
        .prefetch_related("lineas", "lineas__proveedor"),
        pk=pk,
    )

    buffer = BytesIO()

    try:
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        story = []

        wo_num = getattr(getattr(pr.bom, "workorder", None), "numero", "-")

        story.append(Paragraph("Solicitud de Compra", styles["Title"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"<b>OT:</b> {wo_num}", styles["Normal"]))
        story.append(Paragraph(f"<b>BOM:</b> {pr.bom}", styles["Normal"]))
        story.append(Paragraph(f"<b>Estado:</b> {pr.estado}", styles["Normal"]))
        story.append(Spacer(1, 12))

        header = [
            "Plano", "Código", "Descripción", "U/M",
            "Req.", "Disp.", "A comprar",
            "Proveedor", "P.Unit", "Subtotal"
        ]

        data = [header]
        total = Decimal("0")

        cell_style = styles["Normal"]
        cell_style.fontSize = 8
        cell_style.leading = 9

        for ln in pr.lineas.filter(cantidad_requerida__gt=0):
            a = Decimal(ln.cantidad_a_comprar or 0)
            p = Decimal(ln.precio_unitario or 0)
            subtotal = a * p
            total += subtotal

            descripcion = Paragraph(
                (ln.descripcion or "").replace("\n", "<br/>"),
                cell_style
            )

            proveedor = Paragraph(
                (ln.proveedor.nombre if ln.proveedor else ""),
                cell_style
            )

            data.append([
                ln.plano or "",
                ln.codigo or "",
                descripcion,
                ln.unidad or "",
                f"{Decimal(ln.cantidad_requerida or 0):,.3f}",
                f"{Decimal(ln.cantidad_disponible or 0):,.3f}",
                f"{a:,.3f}",
                proveedor,
                f"{p:,.2f}",
                f"{subtotal:,.2f}",
            ])

        data.append(["", "", "", "", "", "", "", "", "TOTAL", f"{total:,.2f}"])

        table = Table(
            data,
            colWidths=[45, 55, 180, 30, 35, 35, 45, 80, 45, 55],
            repeatRows=1,
        )

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))

        story.append(table)
        doc.build(story)

        buffer.seek(0)

        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = (
            f'attachment; filename="solicitud_compra_{pr.pk}.pdf"'
        )
        return resp

    finally:
        buffer.close()


@login_required
def purchase_request_excel(request, pk: int):
    pr = get_object_or_404(
        PurchaseRequest.objects
        .select_related("bom", "bom__workorder")
        .prefetch_related("lineas", "lineas__proveedor"),
        pk=pk,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitud de Compra"

    wo_num = getattr(getattr(pr.bom, "workorder", None), "numero", "-")

    ws.append(["Solicitud de Compra"])
    ws.append(["OT", str(wo_num)])
    ws.append(["BOM", str(pr.bom)])
    ws.append(["Estado", str(pr.estado)])
    ws.append([])

    headers = [
        "Plano", "Código", "Descripción", "U/M",
        "Req.", "Disp.", "A comprar",
        "Proveedor", "P.Unit", "Subtotal"
    ]

    ws.append(headers)

    total = Decimal("0")

    for ln in pr.lineas.filter(cantidad_requerida__gt=0):
        a = Decimal(ln.cantidad_a_comprar or 0)
        p = Decimal(ln.precio_unitario or 0)
        subtotal = a * p
        total += subtotal

        ws.append([
            ln.plano or "",
            ln.codigo or "",
            ln.descripcion or "",
            ln.unidad or "",
            float(Decimal(ln.cantidad_requerida or 0)),
            float(Decimal(ln.cantidad_disponible or 0)),
            float(a),
            ln.proveedor.nombre if ln.proveedor else "",
            float(p),
            float(subtotal),
        ])

    ws.append(["", "", "", "", "", "", "", "", "TOTAL", float(total)])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="solicitud_compra_{pr.pk}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def paw_detail(request, pk):
    compra = get_object_or_404(
        PurchaseRequest.objects
        .select_related("bom", "bom__workorder", "creado_por")
        .prefetch_related(
            "lineas__proveedor",
            "lineas__finance_line",
            "lineas__purchase_approval_line",
        ),
        pk=pk
    )

    from .forms import PurchaseLineFormSet

    queryset = compra.lineas.filter(cantidad_requerida__gt=0).order_by("id")

    if request.method == "POST":
        if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
            messages.error(request, "No tienes permiso para editar esta solicitud.")
            return redirect("compras_oil:paw_detail", pk=compra.pk)

        nuevo_estado = request.POST.get("estado")
        if nuevo_estado in ["BORRADOR", "EN_REVISION"]:
            compra.estado = nuevo_estado
            compra.save(update_fields=["estado", "actualizado_en"])

        formset = PurchaseLineFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Solicitud de compra actualizada correctamente.")
            return redirect("compras_oil:paw_detail", pk=compra.pk)
    else:
        formset = PurchaseLineFormSet(queryset=queryset)

    total_requerido = Decimal("0")
    total_a_comprar = Decimal("0")
    for ln in queryset:
        precio = ln.precio_unitario or Decimal("0")
        total_requerido += (ln.cantidad_requerida or Decimal("0")) * precio
        total_a_comprar += (ln.cantidad_a_comprar or Decimal("0")) * precio

    # Adjunta el estado individual a cada instancia usada por el formset.
    for form in formset.forms:
        info = _estado_flujo_linea(form.instance)
        form.instance.flujo_estado = info["estado"]
        form.instance.flujo_texto = info["texto"]
        form.instance.flujo_aprobado = info["aprobado"]
        form.instance.flujo_enviado = info["enviado"]
        form.instance.en_inventario = _linea_en_inventario(form.instance)
        form.instance.puede_enviar_inventario = (
            info["aprobado"]
            and not form.instance.en_inventario
            and Decimal(form.instance.cantidad_a_comprar or 0) > 0
            and compra.estado != "CERRADA"
        )

    paw = _get_paw_from_compra(compra)
    resumen = _resumen_aprobaciones(compra)
    flujo_recepcion_creada = _recepcion_creada(compra)

    # Recepción activa: permite entrar a Inventario sin volver a ejecutar un envío general.
    try:
        from inventario.models import InventoryReception
        recepcion = InventoryReception.objects.filter(purchase_request=compra).first()
    except Exception:
        recepcion = None

    flujo_recepcion_ok = _recepcion_completa(compra)
    entrega = _get_entrega(compra)
    flujo_entrega_creada = entrega is not None
    flujo_entrega_ok = _entrega_completa(compra)

    lineas_aprobadas_pendientes_inventario = [
        ln for ln in queryset
        if _estado_flujo_linea(ln)["aprobado"]
        and Decimal(ln.cantidad_a_comprar or 0) > 0
        and not _linea_en_inventario(ln)
    ]

    puede_enviar_inventario = (
        bool(lineas_aprobadas_pendientes_inventario)
        and compra.estado != "CERRADA"
    )
    puede_registrar_recepcion = (
        resumen["todas_aprobadas"]
        and flujo_recepcion_creada
        and not flujo_recepcion_ok
        and compra.estado != "CERRADA"
    )
    puede_generar_entrega = flujo_recepcion_ok and not flujo_entrega_creada and compra.estado != "CERRADA"
    puede_cerrar_compra = flujo_recepcion_ok and flujo_entrega_ok and compra.estado != "CERRADA"

    if compra.estado == "CERRADA":
        siguiente_paso = "Compra cerrada"
    elif lineas_aprobadas_pendientes_inventario:
        siguiente_paso = (
            f"Enviar a Inventario los ítems aprobados "
            f"({len(lineas_aprobadas_pendientes_inventario)} pendiente(s) de envío)"
        )
    elif not resumen["todas_aprobadas"]:
        siguiente_paso = f"Completar aprobaciones por ítem ({resumen['aprobadas']}/{resumen['total']})"
    elif not flujo_recepcion_creada:
        siguiente_paso = "Enviar a inventario"
    elif not flujo_recepcion_ok:
        siguiente_paso = "Registrar recepción de material"
    elif not flujo_entrega_creada:
        siguiente_paso = "Definir destino y generar entrega"
    elif not flujo_entrega_ok:
        siguiente_paso = "Completar cantidades entregadas"
    else:
        siguiente_paso = "Cerrar compra"

    return render(request, "compras_oil/paw_detail.html", {
        "compra": compra,
        "lineas": queryset,
        "formset": formset,
        "total_requerido": total_requerido,
        "total_a_comprar": total_a_comprar,
        "puede_compras": tiene_rol(request.user, ["COMPRAS", "ADMIN"]),
        "paw": paw,
        "resumen_aprobaciones": resumen,
        "flujo_recepcion_creada": flujo_recepcion_creada,
        "recepcion": recepcion,
        "flujo_recepcion_ok": flujo_recepcion_ok,
        "flujo_entrega_creada": flujo_entrega_creada,
        "flujo_entrega_ok": flujo_entrega_ok,
        "entrega": entrega,
        "puede_enviar_inventario": puede_enviar_inventario,
        "lineas_aprobadas_pendientes_inventario": lineas_aprobadas_pendientes_inventario,
        "puede_registrar_recepcion": puede_registrar_recepcion,
        "puede_generar_entrega": puede_generar_entrega,
        "puede_cerrar_compra": puede_cerrar_compra,
        "siguiente_paso": siguiente_paso,
    })


@require_POST
@login_required
def enviar_linea_finanzas(request, linea_id):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para enviar ítems a Finanzas.")
        return redirect("/")

    from finanzas.models import FinanceApproval, FinanceApprovalLine

    linea = get_object_or_404(
        PurchaseLine.objects.select_related("request", "proveedor"),
        pk=linea_id,
    )
    compra = linea.request

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if linea.tipo_pago != "CONTADO":
        messages.error(request, "Este ítem no es de contado; debe enviarse a Gerencia.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if Decimal(linea.cantidad_a_comprar or 0) <= 0 or not linea.proveedor_id or linea.precio_unitario is None:
        messages.error(request, "Completa cantidad, proveedor y precio antes de enviar el ítem.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    finanza, _ = FinanceApproval.objects.get_or_create(
        purchase_request=compra,
        defaults={"estado": FinanceApproval.Estado.PENDIENTE, "enviado_por": request.user},
    )
    FinanceApprovalLine.objects.get_or_create(approval=finanza, purchase_line=linea)

    messages.success(request, f"Ítem {linea.codigo or linea.id} enviado a Finanzas.")
    return redirect("compras_oil:paw_detail", pk=compra.pk)


@require_POST
@login_required
def enviar_linea_gerencia(request, linea_id):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para enviar ítems a Gerencia.")
        return redirect("/")

    from aprobacion.models import PurchaseApproval, PurchaseApprovalLine

    linea = get_object_or_404(
        PurchaseLine.objects.select_related("request", "proveedor"),
        pk=linea_id,
    )
    compra = linea.request

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if linea.tipo_pago != "CREDITO":
        messages.error(request, "Este ítem no es de crédito; debe enviarse a Finanzas.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if Decimal(linea.cantidad_a_comprar or 0) <= 0 or not linea.proveedor_id or linea.precio_unitario is None:
        messages.error(request, "Completa cantidad, proveedor y precio antes de enviar el ítem.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    aprobacion, _ = PurchaseApproval.objects.get_or_create(
        purchase_request=compra,
        defaults={"estado": PurchaseApproval.Estado.PENDIENTE, "enviado_por": request.user},
    )
    aprobacion_linea, _ = PurchaseApprovalLine.objects.get_or_create(
        approval=aprobacion,
        purchase_line=linea,
    )
    # Refresca el snapshot mientras la decisión siga pendiente.
    if aprobacion_linea.estado_aprobacion == PurchaseApprovalLine.EstadoAprobacion.PENDIENTE:
        aprobacion_linea.snapshot_from_purchase_line()
        aprobacion_linea.save()

    aprobacion.recalcular_estado()
    aprobacion.save(update_fields=["estado", "actualizado_en"])

    messages.success(request, f"Ítem {linea.codigo or linea.id} enviado a Gerencia.")
    return redirect("compras_oil:paw_detail", pk=compra.pk)


@require_POST
@login_required
def enviar_finanzas(request, pk):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para enviar a Finanzas.")
        return redirect("/")

    from finanzas.models import FinanceApproval, FinanceApprovalLine

    compra = get_object_or_404(
        PurchaseRequest.objects.prefetch_related("lineas"),
        pk=pk
    )

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if not _tiene_lineas_contado(compra):
        messages.error(
            request,
            "No hay líneas de pago contado para enviar a Finanzas."
        )
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    paw = _get_paw_from_compra(compra)

    if paw:
        paw.estado_operativo = "EN_FINANZAS"
        paw.save(update_fields=["estado_operativo"])

    finanza, created = FinanceApproval.objects.get_or_create(
        purchase_request=compra,
        defaults={
            "estado": FinanceApproval.Estado.PENDIENTE,
            "enviado_por": request.user,
        },
    )

    lineas_contado = compra.lineas.filter(
        tipo_pago="CONTADO",
        cantidad_requerida__gt=0,
        cantidad_a_comprar__gt=0,
        proveedor__isnull=False,
    )

    creadas = 0

    for linea in lineas_contado:
        _, linea_creada = FinanceApprovalLine.objects.get_or_create(
            approval=finanza,
            purchase_line=linea,
        )

        if linea_creada:
            creadas += 1

    messages.success(
        request,
        f"PAW enviado a Finanzas correctamente. Líneas nuevas sincronizadas: {creadas}."
    )

    return redirect("compras_oil:paw_detail", pk=compra.pk)


@require_POST
@login_required
def enviar_aprobacion(request, pk):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para enviar a Aprobación.")
        return redirect("/")

    from aprobacion.models import PurchaseApproval
    from aprobacion.admin import sync_purchase_approval_lines

    compra = get_object_or_404(PurchaseRequest.objects.prefetch_related("lineas"), pk=pk)

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    lineas_reales = compra.lineas.filter(
        cantidad_requerida__gt=0,
        cantidad_a_comprar__gt=0,
        proveedor__isnull=False,
        precio_unitario__isnull=False,
    ).exists()

    if not lineas_reales:
        messages.error(
            request,
            "No hay líneas de compra diligenciadas para enviar a Aprobación."
        )
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if _tiene_lineas_contado(compra) and not _finanzas_aprobado(compra):
        messages.error(
            request,
            "Esta compra tiene líneas de contado. Primero debes enviarla a Finanzas y esperar aprobación/pago."
        )
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    paw = _get_paw_from_compra(compra)

    if paw:
        paw.estado_operativo = "EN_APROBACION"
        paw.save(update_fields=["estado_operativo"])

    aprobacion, created = PurchaseApproval.objects.get_or_create(
        purchase_request=compra,
        defaults={
            "estado": PurchaseApproval.Estado.PENDIENTE,
            "enviado_por": request.user,
        },
    )

    sync_purchase_approval_lines(aprobacion, refresh_pending_only=True)
    aprobacion.recalcular_estado()
    aprobacion.save()

    messages.success(
        request,
        "PAW enviado a Aprobación de Compras correctamente."
    )
    return redirect("compras_oil:paw_detail", pk=compra.pk)

@require_POST
@login_required
def enviar_linea_inventario(request, linea_id):
    """Envía a Inventario una sola línea, siempre que ya esté aprobada."""
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para enviar ítems a Inventario.")
        return redirect("/")

    from inventario.models import InventoryReception, InventoryReceptionLine

    linea = get_object_or_404(
        PurchaseLine.objects.select_related("request", "proveedor"),
        pk=linea_id,
    )
    compra = linea.request

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    info = _estado_flujo_linea(linea)
    if not info["aprobado"]:
        messages.error(
            request,
            f"El ítem {linea.codigo or linea.id} todavía no está aprobado y no puede enviarse a Inventario."
        )
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    cantidad = Decimal(linea.cantidad_a_comprar or 0)
    if cantidad <= 0:
        messages.error(request, "Este ítem no tiene cantidad pendiente por comprar.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    recepcion, _ = InventoryReception.objects.get_or_create(
        purchase_request=compra,
        defaults={"creado_por": request.user},
    )

    recepcion_linea, created = InventoryReceptionLine.objects.get_or_create(
        recepcion=recepcion,
        purchase_line=linea,
        defaults={
            "codigo": linea.codigo or "",
            "descripcion": linea.descripcion or "",
            "unidad": linea.unidad or "",
            "cantidad_esperada": cantidad,
            "cantidad_recibida": 0,
            "estado": "PENDIENTE",
        },
    )

    if created:
        messages.success(
            request,
            f"Ítem {linea.codigo or linea.id} enviado a Inventario. Ya puede ser recibido independientemente."
        )
    else:
        messages.info(
            request,
            f"El ítem {linea.codigo or linea.id} ya había sido enviado a Inventario."
        )

    return redirect("compras_oil:paw_detail", pk=compra.pk)


@require_POST
@login_required
def enviar_inventario(request, pk):
    """Envía a Inventario todos los ítems que ya estén aprobados y aún no hayan sido enviados."""
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para enviar a Inventario.")
        return redirect("/")

    from inventario.models import InventoryReception, InventoryReceptionLine

    compra = get_object_or_404(
        PurchaseRequest.objects.prefetch_related("lineas"),
        pk=pk,
    )

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    lineas_aprobadas = []
    for linea in compra.lineas.filter(cantidad_a_comprar__gt=0):
        if _estado_flujo_linea(linea)["aprobado"]:
            lineas_aprobadas.append(linea)

    if not lineas_aprobadas:
        messages.error(request, "No hay ítems aprobados pendientes para enviar a Inventario.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    recepcion, _ = InventoryReception.objects.get_or_create(
        purchase_request=compra,
        defaults={"creado_por": request.user},
    )

    creadas = 0
    existentes = 0
    for ln in lineas_aprobadas:
        cantidad = Decimal(ln.cantidad_a_comprar or 0)
        _, created = InventoryReceptionLine.objects.get_or_create(
            recepcion=recepcion,
            purchase_line=ln,
            defaults={
                "codigo": ln.codigo or "",
                "descripcion": ln.descripcion or "",
                "unidad": ln.unidad or "",
                "cantidad_esperada": cantidad,
                "cantidad_recibida": 0,
                "estado": "PENDIENTE",
            },
        )
        if created:
            creadas += 1
        else:
            existentes += 1

    if creadas:
        messages.success(
            request,
            f"Se enviaron {creadas} ítem(s) aprobado(s) a Inventario. "
            f"Los demás pueden enviarse después cuando sean aprobados."
        )
    else:
        messages.info(request, "Todos los ítems aprobados ya estaban enviados a Inventario.")

    return redirect("inventario:recepcion_detail", pk=recepcion.pk)

@require_POST
@login_required
def generar_entrega(request, pk):
    if not tiene_rol(request.user, ["COMPRAS", "ADMIN"]):
        messages.error(request, "No tienes permiso para generar entregas.")
        return redirect("/")

    from inventario.models import WorkshopDelivery, WorkshopDeliveryLine

    compra = get_object_or_404(PurchaseRequest.objects.prefetch_related("lineas"), pk=pk)

    if compra.estado == "CERRADA":
        messages.error(request, "No puedes modificar una compra cerrada.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    if not _recepcion_completa(compra):
        messages.error(request, "Primero debes completar la recepción del material.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    destino = (request.POST.get("destino") or "").upper().strip()
    destinos_validos = {"TALLER", "CAMPO", "INVENTARIO"}
    if destino not in destinos_validos:
        messages.error(request, "Selecciona un destino válido para la entrega.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    paw = _get_paw_from_compra(compra)
    if destino == "TALLER" and paw and not getattr(paw, "aplica_taller", False):
        messages.error(request, "Este PAW no tiene habilitado el alcance Taller.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)
    if destino == "CAMPO" and paw and not getattr(paw, "aplica_campo", False):
        messages.error(request, "Este PAW no tiene habilitado el alcance Campo.")
        return redirect("compras_oil:paw_detail", pk=compra.pk)

    entrega, created = WorkshopDelivery.objects.get_or_create(
        purchase_request=compra,
        defaults={"creado_por": request.user, "destino": destino},
    )

    if not created and getattr(entrega, "destino", destino) != destino:
        if any(Decimal(x.cantidad_entregada or 0) > 0 for x in entrega.lineas.all()):
            messages.error(request, "No puedes cambiar el destino porque la entrega ya tiene cantidades registradas.")
            return redirect("compras_oil:paw_detail", pk=compra.pk)
        entrega.destino = destino
        entrega.save(update_fields=["destino", "actualizado_en"])

    creadas = 0
    for ln in compra.lineas.filter(cantidad_a_comprar__gt=0):
        cantidad = Decimal(ln.cantidad_a_comprar or 0)
        _, created_line = WorkshopDeliveryLine.objects.get_or_create(
            delivery=entrega,
            purchase_line=ln,
            defaults={
                "codigo": ln.codigo or "",
                "descripcion": ln.descripcion or "",
                "unidad": ln.unidad or "",
                "cantidad_requerida": cantidad,
            },
        )
        if created_line:
            creadas += 1

    messages.success(request, f"Entrega a {entrega.get_destino_display()} generada. Líneas nuevas: {creadas}.")
    return redirect("inventario:entrega_taller_detail", pk=entrega.pk)


# Alias temporal para enlaces antiguos.
generar_entrega_taller = generar_entrega